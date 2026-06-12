#!/usr/bin/env python3
"""
Generate an Excel report of all your GitHub forks with detailed information.
Requires: pip install PyGithub openpyxl
"""

import os
import re
from datetime import datetime
from github import Github
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Category detection keywords
CATEGORY_KEYWORDS = {
    'AI/ML': ['ai', 'agent', 'llm', 'gpt', 'claude', 'machine learning', 'neural', 'deep learning', 'model', 'transformer'],
    'Prompt Engineering': ['prompt', 'prompting', 'gpt', 'chatgpt', 'claude-code'],
    'Coding Tools': ['code', 'editor', 'ide', 'cursor', 'copilot', 'coder', 'dev', 'programming'],
    'Automation': ['automation', 'workflow', 'script', 'bot', 'agent', 'automate'],
    'Organization': ['organize', 'manager', 'task', 'todo', 'calendar', 'file', 'sort', 'tag'],
    'Design': ['design', 'ui', 'ux', 'diagram', 'visual', 'draw', 'figma', 'sketch'],
    'Data': ['data', 'database', 'sql', 'analytics', 'visualization', 'chart', 'graph', 'spreadsheet'],
    'API/Integration': ['api', 'integration', 'webhook', 'proxy', 'connector', 'sync'],
    'Productivity': ['productivity', 'noter', 'markdown', 'obsidian', 'docs', 'notebook'],
    'Remote Work': ['remote', 'job', 'work', 'career'],
    'System': ['system', 'os', 'terminal', 'shell', 'cli', 'command', 'font', 'theme'],
    'Web': ['web', 'browser', 'extension', 'chrome', 'firefox', 'website'],
    'Media': ['music', 'video', 'image', 'photo', 'spotify', 'youtube', 'podcast'],
    'Security': ['security', 'password', 'auth', 'encryption', 'credential'],
    'Email': ['email', 'gmail', 'mail'],
}


def detect_category(repo_name, description):
    """Detect category based on repo name and description."""
    text = f"{repo_name} {description}".lower()
    
    for category, keywords in CATEGORY_KEYWORDS.items():
        for keyword in keywords:
            if keyword in text:
                return category
    
    return 'Other'


def get_forks(username, token=None):
    """Fetch all forked repositories for a user."""
    if token:
        g = Github(token)
    else:
        g = Github()
    
    user = g.get_user(username)
    forks = []
    
    print(f"Fetching forks for {username}...")
    for repo in user.get_repos(type="owner"):
        if repo.fork:
            # Get original repo info
            original_repo = repo.source
            
            category = detect_category(repo.name, repo.description or "")
            
            forks.append({
                'name': repo.name,
                'category': category,
                'fork_url': repo.html_url,
                'original_url': original_repo.html_url,
                'description': repo.description or 'No description',
                'stars': original_repo.stargazers_count,
                'language': repo.language or 'N/A',
                'original_created': original_repo.created_at.strftime('%Y-%m-%d') if original_repo.created_at else 'N/A',
                'fork_created': repo.created_at.strftime('%Y-%m-%d') if repo.created_at else 'N/A',
                'forks_count': original_repo.forks_count,
                'open_issues': original_repo.open_issues_count,
            })
    
    # Sort alphabetically by repository name
    return sorted(forks, key=lambda x: x['name'].lower())


def create_excel(forks, username, output_filename='forks_report.xlsx'):
    """Create an Excel report of forks."""
    
    if not forks:
        print("No forks found!")
        return
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Forks Report"
    
    # Define styles
    header_fill = PatternFill(start_color="0366d6", end_color="0366d6", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    data_font = Font(size=10)
    data_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    center_alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
    
    border = Border(
        left=Side(style='thin', color='d1d5da'),
        right=Side(style='thin', color='d1d5da'),
        top=Side(style='thin', color='d1d5da'),
        bottom=Side(style='thin', color='d1d5da')
    )
    
    # Define headers
    headers = [
        'Repository Name',
        'Category',
        'Your Fork URL',
        'Original Repo URL',
        'Description',
        'Stars (Original)',
        'Language',
        'Original Created',
        'Fork Created',
        'Forks (Original)',
        'Open Issues (Original)'
    ]
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Set column widths
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 12
    ws.column_dimensions['K'].width = 15
    
    # Write data rows
    for row_num, fork in enumerate(forks, 2):
        # Repository Name
        cell = ws.cell(row=row_num, column=1)
        cell.value = fork['name']
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = border
        
        # Category
        cell = ws.cell(row=row_num, column=2)
        cell.value = fork['category']
        cell.font = data_font
        cell.alignment = center_alignment
        cell.border = border
        
        # Your Fork URL (hyperlink)
        cell = ws.cell(row=row_num, column=3)
        cell.value = fork['name']
        cell.hyperlink = fork['fork_url']
        cell.font = Font(color="0366d6", underline="single", size=10)
        cell.alignment = data_alignment
        cell.border = border
        
        # Original Repo URL (hyperlink)
        cell = ws.cell(row=row_num, column=4)
        repo_display = fork['original_url'].split('/')[-1]
        cell.value = repo_display
        cell.hyperlink = fork['original_url']
        cell.font = Font(color="0366d6", underline="single", size=10)
        cell.alignment = data_alignment
        cell.border = border
        
        # Description
        cell = ws.cell(row=row_num, column=5)
        cell.value = fork['description']
        cell.font = data_font
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell.border = border
        
        # Stars
        cell = ws.cell(row=row_num, column=6)
        cell.value = fork['stars']
        cell.font = data_font
        cell.alignment = center_alignment
        cell.border = border
        
        # Language
        cell = ws.cell(row=row_num, column=7)
        cell.value = fork['language']
        cell.font = data_font
        cell.alignment = center_alignment
        cell.border = border
        
        # Original Created
        cell = ws.cell(row=row_num, column=8)
        cell.value = fork['original_created']
        cell.font = data_font
        cell.alignment = center_alignment
        cell.border = border
        
        # Fork Created
        cell = ws.cell(row=row_num, column=9)
        cell.value = fork['fork_created']
        cell.font = data_font
        cell.alignment = center_alignment
        cell.border = border
        
        # Forks Count
        cell = ws.cell(row=row_num, column=10)
        cell.value = fork['forks_count']
        cell.font = data_font
        cell.alignment = center_alignment
        cell.border = border
        
        # Open Issues
        cell = ws.cell(row=row_num, column=11)
        cell.value = fork['open_issues']
        cell.font = data_font
        cell.alignment = center_alignment
        cell.border = border
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Save workbook
    wb.save(output_filename)
    print(f"✅ Excel generated: {output_filename}")


def main():
    """Main function."""
    import sys
    
    # Configuration
    USERNAME = 'EeveeVictoria'
    TOKEN = os.getenv('GITHUB_TOKEN')
    OUTPUT_FILE = 'forks_report.xlsx'
    
    try:
        # Fetch forks
        forks = get_forks(USERNAME, TOKEN)
        
        if not forks:
            print("No forks found!")
            return
        
        print(f"Found {len(forks)} forks")
        
        # Create Excel
        create_excel(forks, USERNAME, OUTPUT_FILE)
        print(f"\n📄 Report saved to: {OUTPUT_FILE}")
        print(f"Generated on {datetime.now().strftime('%B %d, %Y at %I:%M %p')}")
        
    except Exception as e:
        print(f"❌ Error: {e}")
        sys.exit(1)


if __name__ == '__main__':
    main()
